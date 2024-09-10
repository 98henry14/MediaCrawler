import concurrent.futures
import datetime
import logging
import os.path
import sys
import threading

from openpyxl import load_workbook,Workbook
import csv
import xlrd
import requests
from openpyxl.styles import Font, PatternFill, Alignment

import subprocess

def run_azure_command(command):
    """
     执行azure命令
    """
    try:
        result = subprocess.run(command, check=True, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        return result.stdout.strip()
    except subprocess.CalledProcessError as e:
        print(f"Error running command: {e}")
        return None

def split_array(arr, chunk_size=100):
    """
    按指定数目再次分割数组
    """
    # 计算子数组的数量
    num_chunks = (len(arr) + chunk_size - 1) // chunk_size

    # 分割数组
    chunks = [arr[i * chunk_size:(i + 1) * chunk_size] for i in range(num_chunks)]

    return chunks


def call_api(name_list ,queryField,href=None):
    if not name_list:
        exit()

    if href:
        url = href
    else:
        # 需要拼接url,每100条调一次接口查询
        print(f"线程{threading.currentThread().getName()}-正在处理以下信息:{name_list}")
        filter = " or ".join([f"{queryField} eq \"{e}\"" for e in name_list])
        url = f"https://api.ed.pg.com/directory/v1/ou=people,ou=pg,o=world/subtree?searchScope=singleLevel&limit=100&filter={filter}&includeAttributes=uid,extShortName,mail,sn,cn"

    payload = {}
    headers = {
        'Authorization': 'Basic dWlkPUVYNTY3NCxvdT1wZW9wbGUsb3U9cGcsbz13b3JsZDpUc2M3NDUxNTkhamV0',
        'Content-Type': 'application/json'
    }
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        # print(response.json())
        json = response.json()
        entry = json.get("_embedded").get("entries")
        if json.get("_links").get("next"):
            href = json.get("_links").get("next").get("href")
            entry.extend(call_api(name_list, queryField,href))

        return entry


ADD_user_map = {}

def read_add_user_list():
    print(f"{threading.current_thread().getName()} 读取add的用户列表")
    start = datetime.datetime.now()
    # 方式1 ,从之前拉取的azure user文件中匹配,速度快
    # with open("C:\\Users\\xiaoj\\Desktop\\usercheck\\ad_users-2.csv", "r") as f:
    #     rd = csv.reader(f)
    #     for row in rd:
    #         # add_user_list.append(row[0])
    #         ADD_user_map.update({row[0].lower(): row[0] })

    # print(run_azure_command("az cloud set -n AzureChinaCloud"))
    # 方式2:每次都调azure cli 拉取用户信息,至少需要一分钟的处理时间
    cmd = "az ad user list --query [].userPrincipalName --output tsv"
    res = run_azure_command(cmd)
    if res:
        # 写入本地表
        # with open("C:\\Users\\xiaoj\\Desktop\\usercheck\\ad_users-2.csv",mode="w",newline='',encoding="utf-8") as ff:
        #     write = csv.writer(ff)
        #     write.writerows([[line] for line in res.split("\n")])
        with threading.Lock():
            # add_user_list.extend(res.split("\n"))
            ADD_user_map.update({e.lower(): e for e in res.split("\n")})
    print(f"{threading.current_thread().getName()} 读取完毕,耗时：{datetime.datetime.now() - start}，ADD_user_map size:{len(ADD_user_map)}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("please input the action type,like -f [yourfile] ")
        exit()
    if str(sys.argv[1]) not in ['-f']:
        print("can not verify the action type")
        exit()
    filepaths = str(sys.argv[2])
    option_map = {}
    if len(sys.argv) > 3:
        for arg in sys.argv[3:]:
            if arg.__contains__("="):
                kv = arg.split("=")
                option_map.update({kv[0]: kv[1]})
    print("option_map", sys.argv[3:], option_map)
    assign_col = int(option_map.get("colFrom")) if option_map.get("colFrom") else 0
    queryField = option_map.get("queryField") if option_map.get("queryField") else 'extShortName'
    t = threading.Thread(target=read_add_user_list)
    t.start()
    try:
        write_map = {}
        total_list = []
        for filepath in filepaths.split(','):
            if os.path.exists(filepath):
                fs = filepath.split('.')
                filename = fs[0].split("\\")[-1]
                filetype = fs[-1]
                if filetype not in ['csv', 'xls', 'xlsx']:
                    print("file type error,only support csv/xls/xlsx")
                else:
                    queryValList = []
                    if filetype == 'csv':
                        print("csv格式处理")
                        with open(filepath, mode='r', encoding='utf-8') as file:
                            reader = csv.reader(file)
                            next(reader)
                            for row in reader:
                                # print(row[0])
                                queryValList.append(row[assign_col])
                    else:
                        print("excel格式处理")
                        if filetype == 'xls':
                            workbook = xlrd.open_workbook(filepath, formatting_info=True)
                        elif filetype == 'xlsx':
                            workbook = xlrd.open_workbook(filepath)

                        sheet = workbook.sheet_by_index(0)
                        for row in range(1, sheet.nrows):
                            # print(sheet.row_values(row)[0])
                            queryValList.append(sheet.row_values(row)[assign_col])
                    write_map[filename] = queryValList
                    total_list.extend(queryValList)
            else:
                print(f"file not exist->{filepath}")
        # 调接口查询,开启多线程
        split_arr = split_array(total_list, 100)
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(call_api, name_list, queryField) for name_list in split_arr]
            results = []
            for future in concurrent.futures.as_completed(futures):
                results.extend(future.result())

        user_map = {item.get(queryField)[0].lower() if queryField in ['uid'] else item.get(queryField).lower(): item for
                    item in results}
        t.join()
        print("等待线程执行后获取到 add用户列表数量:",len(ADD_user_map))
        #     写excel
        for key in write_map.keys():
            queryValList = write_map.get(key)
            wfile = f"./result-{key}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            headers = [f"queryfield-{queryField}", "uid","mail", "exists in ED","exists in ADD"]
            ws.append(headers)
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')


            for index,name in enumerate(queryValList):
                # 对应的的行涉及表头跟list索引的起始位为0，需要+2
                row = index + 2
                # 转为小写匹配
                ln = name.lower()
                uid = user_map.get(ln).get("uid")[0] if user_map.get(ln) and user_map.get(ln).get("uid") else ""
                mail = user_map.get(ln).get("mail")[0] if user_map.get(ln) and user_map.get(ln).get("mail") else ""
                exists_ed = user_map.get(ln) is not None
                exists_add = ADD_user_map.get(user_map.get(ln).get("mail")[0].lower()) is not None if user_map.get(ln) and user_map.get(ln).get("mail") else False

                ws.append([name,uid, mail, exists_ed, exists_add])

                if not exists_ed:
                    cell = ws.cell(row=row, column=4)
                    # 如果不存在则设置为黄色背景
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if not exists_add:
                    cell = ws.cell(row=row, column=5)
                    # 如果不存在则设置为黄色背景
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            wb.save(wfile)

    except Exception as e:
        logging.exception(e)
        print(e)



